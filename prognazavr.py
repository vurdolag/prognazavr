# -*- coding: cp1251 -*-
import requests
import ujson as json
import re
from textdistance import damerau_levenshtein as dist
import io
from PIL import Image, ImageDraw, ImageFont, ImageEnhance
import time
import pickle
import os
import sys
import traceback
import openpyxl

command = ["дедал", "атом", "поволжье", "звезда", "комета", "пента", "оияи",
           "криптен", "промтех", "феникс", "кураж", "камаз", "атлетик"]

path_battle_of_bet_data = 'bd/battle_of_bet/data/'
path_battle_of_bet_result = 'bd/battle_of_bet/result/math_results'
path_best_player_data = 'bd/best_player/data/'
path_best_player_result = 'bd/best_player/result/math_results'

try:
    with open('token.txt') as f:
        serv_token = f.readline()
except FileNotFoundError:
    serv_token = input('Нету вк токена\nВведи токен\n>>> ')
    with open('token.txt', 'w') as f:
        f.write(serv_token)

try:
    with open('config.json', encoding='utf-8') as f:
        config = json.load(f)
except FileNotFoundError:
    serv_token = print('Нету конфиг файла, создаю дефолтные значения')
    d = {"1": 5, "2": 3, "3": 2, "4": 0}
    print(d)
    with open('config.json', 'w', encoding='utf-8') as f:
        json.dump(d, f)

animation = "|/-\\"


def my_print(res):
    print("#" * 37)
    for i in res:
        if len(i) == 4:
            print('#   ', i[0], ' ' * (10 - len(i[0])), i[1], ' ' * (10 - len(i[1])), i[2], i[3], '   #')
        elif len(i) == 5:
            print('#   ', i[0], ' ' * (10 - len(i[0])), i[1], ' ' * (10 - len(i[1])), i[2], i[3], i[4], '   #')
        elif len(i) == 3:
            print('#   ', i[0], ' ' * (10 - len(i[0])), i[1], ' ' * (10 - len(i[1])), i[2], '   #')
        else:
            print('МОЯ НЕ МОЧЬ РАСПЯЧАТАТЬ')

    print("#" * 37)


def distance(a, b):
    size = len(a) + len(b)
    y = round(((size - dist(a, b)) / size) * 100, 4)
    return y


def print_anim(name, index):
    sys.stdout.write("\r" + name + ('|' * index) + animation[index % len(animation)])
    sys.stdout.flush()


def logs(strings='', name='logs_error.txt'):
    """
    Логер и отлов ошибок, печатает полный трейсбек
    :param strings:
    :param name:
    """
    with open('bd/' + name, 'a', encoding='utf-8') as f:
        if strings:
            f.write(f'{time.strftime("%y-%m-%d %H:%M:%S")} => {strings}')
            f.write('\n>>> ')
        else:
            f.write(f'{time.strftime("%y-%m-%d %H:%M:%S")} ERROR => ')
            a = sys.exc_info()
            traceback.print_exception(*a)
            traceback.print_exception(*a, file=f)
            f.write('\n>>> ')


def check_input() -> bool:
    string = input('да [1] / нет [0] >>> ').lower()
    return 'да' in string or '1' in string


class Statistic:
    __slots__ = 'token'

    def __init__(self, token):
        self.token = token

    def get_comment_from_post(self, post: list):
        print('Получаем комментари', 'wall' + '_'.join(post))

        params = {'owner_id': post[0],
                  'post_id': int(post[1]),
                  'need_likes': 1,
                  'extended': 1,
                  'count': 100,
                  'offset': 0,
                  'thread_items_count': 10,
                  'v': '5.103',
                  'access_token': self.token}

        res = json.loads(requests.get('https://api.vk.com/method/wall.getComments', params=params).content)
        if res.get('error', 0):
            raise Exception(res['error'])
        return res['response']['items']

    def draw_outline(self, draw, x, y, finalText, shadow_color, fnt, align, spacing):

        draw.text(
            (x - 1, y), finalText, fill=shadow_color, font=fnt, align=align, spacing=spacing
        )

        draw.text(
            (x + 1, y), finalText, fill=shadow_color, font=fnt, align=align, spacing=spacing
        )

        draw.text(
            (x, y - 1), finalText, fill=shadow_color, font=fnt, align=align, spacing=spacing
        )

        draw.text(
            (x, y + 1), finalText, fill=shadow_color, font=fnt, align=align, spacing=spacing
        )

        draw.text(
            (x - 1, y - 1),
            finalText,
            fill=shadow_color,
            font=fnt,
            align=align,
            spacing=spacing,
        )

        draw.text(
            (x + 1, y - 1),
            finalText,
            fill=shadow_color,
            font=fnt,
            align=align,
            spacing=spacing,
        )

        draw.text(
            (x - 1, y + 1),
            finalText,
            fill=shadow_color,
            font=fnt,
            align=align,
            spacing=spacing,
        )

        draw.text(
            (x + 1, y + 1),
            finalText,
            fill=shadow_color,
            font=fnt,
            align=align,
            spacing=spacing,
        )

    def draw_text(self, draw, coor, text, fnt):
        self.draw_outline(draw, coor[0], coor[1],
                          text, "#000000", fnt, "center", int(65 / 4))
        draw.text(coor, text, fill="#ffffff", font=fnt,
                  align="center", spacing=int(65 / 4))

    def get_avatar_and_name(self, stat):
        task = []

        for j, x in enumerate(stat):
            time.sleep(0.4)

            y = x[0]
            n = ''

            if y > 200000000000:
                y -= 200000000000
                n = '+'

            if y > 0:
                params = {'user_ids': y,
                          'fields': 'photo_200_orig',
                          'v': '5.103',
                          'access_token': self.token,
                          'lang': 'ru'
                          }
                response = json.loads(requests.get('https://api.vk.com/method/users.get', params).content)['response'][0]
                url_photo = response.get('photo_200_orig', '')
                avatar = io.BytesIO(requests.get(url_photo).content)
                task.append([x[0], str(x[1]), response.get('first_name', '') + n, avatar])

            else:
                img = Image.open('temp.jpg')
                stream = io.BytesIO()
                img.save(stream, format="JPEG", quality=75)
                stream.seek(0)
                im = stream.read()
                avatar = io.BytesIO(im)
                task.append([x[0], str(x[1]), 'Игорь', avatar])

            print_anim('Получаю аватары ', j)

        print('| ok')

        return task

    def add_corners(self, im, rad):
        circle = Image.new("L", (rad * 4, rad * 4), 0)
        draw = ImageDraw.Draw(circle)
        draw.ellipse((2, 2, rad * 4 - 2, rad * 4 - 2), fill=255)
        del draw
        circle = circle.resize((rad * 2, rad * 2), Image.ANTIALIAS)
        alpha = Image.new("L", im.size, 255)
        draw = ImageDraw.Draw(alpha)
        w, h = im.size
        alpha.paste(circle.crop((0, 0, rad, rad)), (0, 0))
        alpha.paste(circle.crop((0, rad, rad, rad * 2)), (0, h - rad))
        alpha.paste(circle.crop((rad, 0, rad * 2, rad)), (w - rad, 0))
        alpha.paste(circle.crop((rad, rad, rad * 2, rad * 2)), (w - rad, h - rad))
        draw.rectangle([(0, 0), (alpha.size[0] - 1, alpha.size[1] - 1)], outline=0)
        del draw

        im.putalpha(alpha)
        im = im.convert('RGBA')
        _, _, _, opacity = (0, 0, 0, 200)
        alpha = im.split()[3]
        alpha = ImageEnhance.Brightness(alpha).enhance(opacity / 255)

        im.putalpha(alpha)
        return im

    def sorter(self, point, date):
        min_date = min(date.values()) - 1
        # сортировка по количеству очков и дате комментария, чем раньше комментарий тем выше в списке
        return list(sorted(list(point.items()), key=lambda x: x[1] + (1 / (date[x[0]] - min_date)), reverse=True))

    def create_image(self, stat, name_out, name_input):
        print('Создаю картинку')
        data = stat[:10]

        name_out = 'result/' + name_out

        img = Image.open(name_input)
        draw = ImageDraw.Draw(img)
        fnt = ImageFont.truetype('GothaProBol.otf', 24)
        avatar = self.get_avatar_and_name(data)
        cr = 0

        if len(avatar) != len(stat[:10]):
            raise Exception("Ошибка получения аватаров пользователей, проверь интернет или vk token")

        for j, i in enumerate(avatar[:10]):
            txt = f'{j + 1}. {i[2]} - {i[1]}'

            av = Image.open(i[3])
            av = av.resize((60, 60), Image.ANTIALIAS)
            av = self.add_corners(av, 10)
            img.paste(av, (50, j * 70 + 40), av)
            self.draw_text(draw, (50 + 70, j * 70 + 60), txt, fnt)
            cr += 1
            print_anim('Рендю поля ', j)
            time.sleep(0.05)

        img = img.crop((0, 0, img.size[0], 70 * cr + 50))

        img.save(name_out + '.jpg', format="JPEG", quality=90)

        print("| Готово ->", name_out + '.jpg')

    def write_xls(self, point, namber_tour, name):
        point = dict(point)
        data = {}

        for j,  user_id in enumerate(point.keys()):
            time.sleep(0.05)

            if user_id < 0:
                continue

            y = user_id
            n = ''
            if y > 200000000000:
                y -= 200000000000
                n = '+'

            params = {'user_ids': y,
                      'v': '5.103',
                      'access_token': self.token,
                      'lang': 'ru'}

            resp = json.loads(requests.get('https://api.vk.com/method/users.get', params).content)

            user = resp.get('response', '')

            if not user:
                print('Error bad request get in xsl', resp, params)
                logs(f'error req get xls {str(resp)} {str(params)}')
                continue

            user = user[0]

            data[user_id] = [point[user_id], user['first_name']+n, user['last_name']+n]

            print_anim('Получаю имена ', j)

        print('| ok')

        for j, i in enumerate(point.items()):
            if i[0] < 0:
                data[i[0]] = [i[1], f'От группы {j}', str(i[0])]

        wb = openpyxl.load_workbook(name)
        l = wb['Лист1']
        flag = 1

        for j, user in enumerate(data.items()):
            v_user = user[1]
            i = 0
            while i < 50:
                target_cell = l['A' + str(i+2)]

                if target_cell.value == f'{v_user[1]} {v_user[2]}':
                    cell = l.cell(row=i+2, column=namber_tour+1)
                    if cell.value and flag:
                        print(f'Колонка тура №{namber_tour} содержит информацию, перезаписать?')
                        if check_input():
                            flag = 0

                        else:
                            print('Запись xls отменена')
                            return False

                    cell.value = v_user[0]
                    break

                if not target_cell.value:
                    target_cell.value = f'{v_user[1]} {v_user[2]}'
                    cell = l.cell(row=i + 2, column=namber_tour + 1)
                    if cell.value and flag:
                        print(f'Колонка тура №{namber_tour} содержит информацию, перезаписать?')
                        if check_input():
                            flag = 0

                        else:
                            print('Запись xls отменена')
                            return False

                    cell.value = v_user[0]
                    break

                else:
                    i += 1

            print_anim('xls запись! ', j)
            time.sleep(0.05)

        wb.save(name)
        print(' xls готово ->', name)


class Game:
    def check(self, pars, index: tuple, commands):
        if not isinstance(pars, list):
            pars = [pars]
        for ind, x in enumerate(pars[index[0]:index[1]]):
            for word in commands:
                if distance(x, word) >= 73:
                    pars[ind] = word
                if 'кураж' in x:
                    pars[ind] = 'кураж'

    def check_command(self, pars, commands):
        if not isinstance(pars, list):
            pars = [pars]
        count = 0
        for x in pars:
            if x not in commands:
                print('Упс... проверь правильность ввода ->', x)
                count += 1

        return True if not count else False

    def RE(self, text):
        text = re.sub(r'\uFF09', ')', text)
        text = re.sub(r'\uFF08', '(', text)
        return re.findall(r'.+?\(.+?\)', text)

    def comment(self, i, text, user_id):
        if i['thread']['count'] != 0:
            t = i['thread']['items'][0]
            if str(user_id) in t['text'] or t['from_id'] == user_id:
                text += t['text']
        return text



class BestPlayer(Game):
    def inputs(self):
        math_results = []
        print("\nВвод результата, формат:\nФамилия Команда\nзакончить - стоп")
        best = input('Лучший игрок:\n>>> ')
        ind = 0
        while True:
            if ind == 0:
                c = best + ' 1'
            else:
                c = input(">>> ").lower() + ' 0'
            if 'стоп' in c:
                break

            if 'сброс' in c:
                math_results = []
                print('Список результатов сброшен')
                continue

            pars = re.findall(r'\w+', c.lower())

            if len(pars) != 3:
                print('Упс... проверь правильность ввода ->', c)
                continue

            self.check(pars, command)

            if self.check_command(pars[1:2]):
                math_results.append(pars)
                ind += 1

        return math_results

    def pars_comment(self, data_comment, result, commands):
        print("Получаем результаты из комментариев")
        itog = {}
        date = {}
        index = 0

        fio = [x[0] for x in result]
        fcomand = [x[1] for x in result]
        best = [x for x in result if x[2] == '1']
        for j, i in enumerate(data_comment):
            user_id = i.get('from_id', '')
            if not user_id:
                continue

            text = i['text']

            if user_id < 0:
                user_id = int(f'{user_id}{j}')

            text = self.comment(i, text, user_id)

            r = self.RE(text)[0]
            if not r:
                print('ошибка парсинга коментария ->', i)
                continue

            h = re.sub(r'\).*', '', r[0]).lower()
            pars = h.split('(')

            pars = [x.strip() for x in pars]

            if len(pars) != 2:
                print('ошибка парсинга коментария ->', i)
                continue

            for ind, x in enumerate(pars[1:2]):
                for word in commands:
                    if distance(x, word) >= 73:
                        pars[ind + 1] = word

                    if 'кураж' in x:
                        pars[ind + 1] = 'кураж'

            if pars[1] not in commands:
                print('Упс... некоректное название команды...', pars)
                continue

            index += 1

            if itog.get(user_id, 0):
                user_id += 200000000000

            itog[user_id] = [pars]
            date[user_id] = i['date']

        point = {}

        log = ''

        try:
            log = f'{time.strftime("%y-%m-%d %H:%M:%S")}\nвсе номинанты: {fio}\n' \
                  f'все команды: {fcomand}\nлучший: {best[0]}\n\n'
            log += '-' * 147 + '\n'

        except:
            logs()

        for namber, i in enumerate(itog.items()):
            namber += 1
            stop = 0
            x = i[1][0]
            for j in result:
                if j[0] in x[0] and x[1] == j[1] and j[2] == '1':
                    point[i[0]] = point.get(i[0], 0) + config['1']
                    log += f'{namber: ^3} | коммент {str(x)[:35]: ^35}| цель {str(j)[:35]: ^35}| очки {5: ^3}|' \
                           f'{" " * 40}|\n'
                    stop = 1
                    break

                elif x[1] == j[1] and j[2] == '1':
                    point[i[0]] = point.get(i[0], 0) + config['3']
                    log += f'{namber: ^3} | коммент {str(x)[:35]: ^35}| цель {str(j)[:35]: ^35}| очки {2: ^3}|' \
                           f'{" " * 40}|\n'
                    stop = 1
                    break

            if stop:
                continue

            for j in result:
                if j[0] in x[0]:
                    point[i[0]] = point.get(i[0], 0) + config['2']
                    log += f'{namber: ^3} | коммент {str(x)[:35]: ^35}| цель {str(j)[:35]: ^35}| очки {3: ^3}' \
                           f'| нашел среди номинантов {str(j[0])[:16]: ^16}|\n'
                    stop = 1
                    break

            if stop:
                continue

            point[i[0]] = point.get(i[0], 0) + config['4']
            xnxn = 'ничего не нашел =('
            log += f'{namber: ^3} | коммент {str(x)[:35]: ^35}| цель {xnxn: ^35}| очки {0: ^3}|' \
                   f'{" " * 40}|\n'

        log += '-' * 147 + '\n\n\n'

        with open('log.txt', 'a', encoding='utf-8') as f:
            f.write(log)

        return point, date, itog


class BattleOfBet(Game):
    def inputs(self):
        math_results = []
        print("\nВвод результата, формат:\nкоманда1 команда2 результат1 результат2\nзакончить - стоп")
        while True:
            c = input(">>> ").lower()
            if 'стоп' in c:
                break

            if 'сброс' in c:
                math_results = []
                print('Список результатов сброшен')
                continue

            pars = re.findall(r'\w+', c.lower())

            if len(pars) != 4:
                print('Упс... проверь правильность ввода ->', c)
                continue

            self.check(pars, (0, 2), command)

            if self.check_command(pars[:2], command):
                math_results.append(pars)

        return math_results

    def pars_comment(self, data_comment, result, commands):
        print("Получаем результаты из комментариев")
        itog = {}
        date = {}
        for j, i in enumerate(data_comment):
            user_id = i['from_id']
            text = i['text']

            if user_id < 0:
                user_id = int(f'{user_id}{j}')

            if user_id in itog:
                continue

            text = self.comment(i, text, user_id)

            r = re.findall(r'(\w+?.+?\w+?.+?\d{1,2}.+?\d{1,2})+?', text)
            temp = []
            for j in r:
                pars = re.findall(r'\w+', j.lower())

                self.check(pars, (0, 2), commands)
                if self.check_command(pars[:2], commands):
                    temp.append(pars)

            itog[user_id] = temp
            date[user_id] = i['date']

        point = {}

        for i in itog.items():
            for x in i[1]:
                for j in result:
                    if x[0] == j[1] and x[1] == j[0]:
                        j = [j[1], j[0], j[3], j[2]]

                    if x[0] == j[0] and x[1] == j[1]:
                        if x[2] == j[2] and x[3] == j[3]:
                            point[i[0]] = point.get(i[0], 0) + config['1']
                            # print(j, x, 5)
                            continue

                        win = (int(x[2]) < int(x[3]) and int(j[2]) < int(j[3])) or (
                                int(x[2]) > int(x[3]) and int(j[2]) > int(j[3]))

                        if int(x[2]) - int(x[3]) == int(j[2]) - int(j[3]) and win or (
                                x[2] == x[3] and j[2] == j[3]):
                            point[i[0]] = point.get(i[0], 0) + config['2']
                            # print(j, x, 3)
                            continue

                        if win:
                            point[i[0]] = point.get(i[0], 0) + config['3']
                            # print(j, x, 1)
                            continue

                        point[i[0]] = point.get(i[0], 0) + config['4']

        return point, date, itog

    def inputs_playof(self):
        math_results = []
        print("\nВвод результата, формат:\nкоманда1 команда2 результат1 результат2 пенальти\nзакончить - стоп")
        while True:
            c = input(">>> ").lower()
            if 'стоп' in c:
                break

            if 'сброс' in c:
                math_results = []
                print('Список результатов сброшен')
                continue

            pars = re.findall(r'\w+', c.lower())

            if len(pars) < 4 or len(pars) > 5:
                print('Упс... проверь правильность ввода ->', c)
                continue

            self.check(pars, (0, 2), command)

            if len(pars) == 5:
                self.check(pars, (4, 5), command)
                if not self.check_command(pars[-1], command):
                    continue

            if self.check_command(pars[:2], command):
                math_results.append(pars)

        return math_results

    def pars_comment_playof(self, data_comment, result, commands):
        print("Получаем результаты из комментариев")
        itog = {}
        date = {}
        for j, i in enumerate(data_comment):
            user_id = i['from_id']
            text = i['text']

            if user_id < 0:
                user_id = int(f'{user_id}{j}')

            if user_id in itog:
                continue

            text = self.comment(i, text, user_id)

            # команда команда 1 1 (командв)
            r = re.findall(r'(\w+?.+?\w+?.+?\d{1,2}.+?\d{1,2}(\(.+?\w+?\)|))+?', text)
            temp = []
            for j in r:
                j = self.RE(j)[0]
                pars = re.findall(r'\w+', j.lower())

                self.check(pars, (0, 2), commands)

                if len(pars) == 5:
                    self.check(pars, (4, 5), commands)
                    if not self.check_command(pars[-1], commands):
                        continue

                if self.check_command(pars[:2], commands):
                    temp.append(pars)

            itog[user_id] = temp
            date[user_id] = i['date']

        point = {}

        for i in itog.items():
            for x in i[1]:
                for j in result:
                    if x[0] == j[1] and x[1] == j[0]:
                        j = [j[1], j[0], j[3], j[2]]

                    if x[0] == j[0] and x[1] == j[1]:
                        if x[2] == j[2] and x[3] == j[3]:
                            point[i[0]] = point.get(i[0], 0) + config['1']
                            # print(j, x, 5)
                            continue

                        win = (int(x[2]) < int(x[3]) and int(j[2]) < int(j[3])) or (
                                int(x[2]) > int(x[3]) and int(j[2]) > int(j[3]))

                        if int(x[2]) - int(x[3]) == int(j[2]) - int(j[3]) and win or (
                                x[2] == x[3] and j[2] == j[3]):
                            point[i[0]] = point.get(i[0], 0) + config['2']
                            # print(j, x, 3)
                            continue

                        if win:
                            point[i[0]] = point.get(i[0], 0) + config['3']
                            # print(j, x, 1)
                            continue

                        point[i[0]] = point.get(i[0], 0) + config['4']

        return point, date, itog


class Worker:
    def __init__(self, games, path, commands, name_xls, token):
        self.path_img = path + 'source.jpg'
        self.path = path
        self.name_xls = name_xls
        self.token = token
        self.commands = commands
        self.mode = ''

        game = games()

        if self.check_mode():
            self.mode = 'reiting/'
            self.data_input = game.inputs
            self.pars_comment = game.pars_comment
        else:
            self.mode = 'playof/'
            self.data_input = game.inputs_playof
            self.pars_comment = game.pars_comment_playof

    def open_data_result(self, post):
        math_results = []
        try:
            with open(f'{self.path}result/{self.mode}math_results{"_".join(post)}', 'rb') as f:
                math_results = pickle.load(f)
        except FileNotFoundError:
            pass

        return math_results

    def save_data_result(self, result, post):
        if not isinstance(post, list):
            post = [post]
        with open(f'{self.path}result/{self.mode}math_results{"_".join(post)}', 'wb') as f:
            pickle.dump(result, f)

    def open_data_point(self, post):
        with open(f'{self.path}data/{self.mode}{post}', 'rb') as f:
            point = pickle.load(f)
        return point

    def save_data_point(self, point, post):
        with open(f'{self.path}data/{self.mode}{post}', 'wb') as f:
            pickle.dump(point, f)

    def create_full_stat(self, file):
        print('Составить полную статистику по всем постам?')
        if check_input():
            all_stat = {}

            for i in file:
                with open(self.path + 'data/' + i, 'rb') as f:
                    stat = pickle.load(f)

                for j in stat:
                    all_stat[j[0]] = all_stat.get(j[0], 0) + j[1]

            stat = list(sorted(list(all_stat.items()), key=lambda x: x[1], reverse=True))

            do = Statistic(self.token)
            do.create_image(stat, 'all_stat', self.path_img)

    def get_input(self, post):
        result = self.data_input()

        while True:
            my_print(result)

            print('Всё правильно?')
            if check_input():
                self.save_data_result(result, post)
                break

            else:
                print('Ну чтож, повторим...')
                result = self.data_input()

        return result

    def start(self, file):
        flag = True
        result = []
        if len(file) != 0:
            print('Инфо по постам в базе:\n')
            for ind, i in enumerate(file):
                print(f'[{ind + 1}] {i}')
            print('\n' + "#" * 45)
            print('Номер поста или ссылка на новый')
            if len(file) > 1:
                print('Составить полную статистику - [0]')
            a = input('>>> ').lower().strip()

            if a == '0':
                self.create_full_stat(file)
                print('PROG OK')
                print('#' * 45 + '\n')
                time.sleep(1)
                return '', '', '', True

            try:
                post = file[int(a) - 1].split('_')
                result = self.open_data_result(post)
                flag = False

            except:
                if 'wall' in a:
                    post = a[a.find('wall') + 4:].split('_')
                else:
                    raise Exception('Ошибка! Некоректная ссылка!')

        else:
            a = input('Ссылка на пост\n>>> ').lower().strip()
            if 'wall' in a:
                post = a[a.find('wall') + 4:].split('_')
            else:
                raise Exception('Ошибка! Некоректная ссылка!')

        return post, flag, result, False

    def write_xls(self, do, point, name):
        print('Записать в xls?')
        if check_input():
            while True:
                try:
                    namber_tour = int(input('Номер тура (число)\n>>> ').lower())
                    do.write_xls(point, namber_tour, name)
                    break

                except Exception as ex:
                    logs()
                    print('Это не число!', ex)
        else:
            return

    def check_mode(self):
        string = input("Рейтинг [0] / Плейоф [1] >>> ").lower()
        return 'рейтинг' in string or '0' in string


    def mainapp(self):
        file = os.listdir(f'{self.path}data/{self.mode}')

        post, flag, result, end = self.start(file)
        if end:
            return True

        p = '_'.join(post)

        if p in file and flag:
            print("В базе уже есть информация по этому посту\nОбработать еще раз?")
            if check_input():
                res = self.open_data_result(post)
                if not res:
                    print('Прошлый результат не обнаружен.')
                    res = self.get_input(post)

                print('Прошлый результат верен?')
                my_print(res)

                if check_input():
                    result = res
                else:
                    result = self.get_input(post)

            else:
                if len(file) > 1:
                    self.create_full_stat(file)
                print('PROG OK')
                print('#' * 45 + '\n')
                time.sleep(1)
                return True

        if not flag:
            do = Statistic(self.token)
            point = self.open_data_point(p)
            self.write_xls(do, point, self.name_xls)
            print('Сделать картинку?')
            if check_input():
                do.create_image(point, p, self.path_img)
            print('PROG OK')
            print('#' * 45 + '\n')
            time.sleep(1)
            return True

        if p not in file:
            result = self.get_input(post)

        try:
            do = Statistic(self.token)

            data_comment = do.get_comment_from_post(post)
            point, date, _ = self.pars_comment(data_comment, result, self.commands)

            point = do.sorter(point, date)

            self.save_data_point(point, p)

            self.write_xls(do, point, self.name_xls)

            print('Сделать картинку?')
            if check_input():
                do.create_image(point, '_'.join(post), self.path_img)

            if len(file) > 1:
                self.create_full_stat(file)

        except Exception as ex:
            logs()
            print('error', ex)

        print('PROG OK')
        print('#' * 45 + '\n')
        time.sleep(1)
        return True


while True:
    print('[1] The Battle of bet\n'
          '[2] The Best player\n'
          '[0] Exit')
    c = input('>>> ')
    if c == '1':

        go = Worker(BattleOfBet,
                    'bd/battle_of_bet/',
                    command,
                    'test.xlsx',
                    serv_token)
        go.mainapp()

    elif c == '2':
        go = Worker(BestPlayer,
                    'bd/best_player/',
                    command,
                    'best_player.xlsx',
                    serv_token)
        go.mainapp()

    elif c == '0':
        print('Пока =) ')
        time.sleep(3)
        exit()

    else:
        print('Моя твоя не понимай')


